"""
Complyo Report Generation System - PDF and Excel Reports
Professional compliance reports with templates and customization
"""

import asyncio
import io
import json
import base64
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any, Union, BinaryIO
from dataclasses import dataclass, asdict
import uuid
import os
import tempfile
import logging

# PDF generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image
from reportlab.platypus.flowables import HRFlowable
from reportlab.graphics.shapes import Drawing, Circle, Rect
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics import renderPDF

# Excel generation  
import openpyxl
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
from openpyxl.chart import PieChart, BarChart, LineChart, Reference
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils.dataframe import dataframe_to_rows

# Template system
from jinja2 import Environment, DictLoader, Template

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

@dataclass
class ReportConfig:
    """Report generation configuration"""
    title: str
    subtitle: Optional[str] = None
    company_name: str = "Complyo"
    company_logo_path: Optional[str] = None
    author: str = "Complyo Platform"
    subject: str = "Website Compliance Report"
    keywords: List[str] = None
    include_charts: bool = True
    include_technical_details: bool = True
    include_recommendations: bool = True
    language: str = "de"
    
    def __post_init__(self):
        if self.keywords is None:
            self.keywords = ["compliance", "DSGVO", "website", "legal"]

@dataclass
class ReportData:
    """Data structure for report generation"""
    scan_results: Dict[str, Any]
    user_info: Dict[str, Any]
    website_info: Dict[str, Any]
    compliance_metrics: Dict[str, Any]
    issues: List[Dict[str, Any]]
    recommendations: List[Dict[str, Any]]
    technical_analysis: Optional[Dict[str, Any]] = None
    monitoring_data: Optional[Dict[str, Any]] = None
    generated_at: datetime = None
    
    def __post_init__(self):
        if self.generated_at is None:
            self.generated_at = datetime.now()

class PDFReportGenerator:
    """Professional PDF report generator"""
    
    def __init__(self, config: ReportConfig):
        """Initialize PDF generator with configuration"""
        
        self.config = config
        self.styles = getSampleStyleSheet()
        self.buffer = io.BytesIO()
        
        # Define custom styles
        self._setup_custom_styles()
        
    def _setup_custom_styles(self):
        """Setup custom paragraph styles"""
        
        # Title style
        self.styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            alignment=1,  # Center
            textColor=colors.HexColor('#2C3E50')
        ))
        
        # Subtitle style
        self.styles.add(ParagraphStyle(
            name='CustomSubtitle',
            parent=self.styles['Heading2'],
            fontSize=16,
            spaceAfter=20,
            alignment=1,  # Center
            textColor=colors.HexColor('#34495E')
        ))
        
        # Section header style
        self.styles.add(ParagraphStyle(
            name='SectionHeader',
            parent=self.styles['Heading2'],
            fontSize=14,
            spaceBefore=20,
            spaceAfter=12,
            textColor=colors.HexColor('#E74C3C'),
            borderWidth=1,
            borderColor=colors.HexColor('#E74C3C'),
            borderPadding=5
        ))
        
        # Issue style (critical)
        self.styles.add(ParagraphStyle(
            name='CriticalIssue',
            parent=self.styles['Normal'],
            fontSize=11,
            textColor=colors.HexColor('#E74C3C'),
            backColor=colors.HexColor('#FADBD8'),
            borderWidth=1,
            borderColor=colors.HexColor('#E74C3C'),
            borderPadding=8
        ))
        
        # Recommendation style
        self.styles.add(ParagraphStyle(
            name='Recommendation',
            parent=self.styles['Normal'],
            fontSize=11,
            textColor=colors.HexColor('#27AE60'),
            backColor=colors.HexColor('#D5F4E6'),
            borderWidth=1,
            borderColor=colors.HexColor('#27AE60'),
            borderPadding=8
        ))
    
    async def generate_compliance_report(self, report_data: ReportData) -> bytes:
        """Generate comprehensive compliance report PDF"""
        
        try:
            # Create document
            doc = SimpleDocTemplate(
                self.buffer,
                pagesize=A4,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=18
            )
            
            # Build report content
            story = []
            
            # Title page
            story.extend(self._create_title_page(report_data))
            story.append(PageBreak())
            
            # Executive summary
            story.extend(self._create_executive_summary(report_data))
            story.append(PageBreak())
            
            # Compliance overview
            story.extend(self._create_compliance_overview(report_data))
            
            # Detailed issues
            story.extend(self._create_issues_section(report_data))
            
            # Technical analysis (if requested)
            if self.config.include_technical_details and report_data.technical_analysis:
                story.extend(self._create_technical_section(report_data))
            
            # Recommendations (if requested)
            if self.config.include_recommendations:
                story.extend(self._create_recommendations_section(report_data))
            
            # Appendix
            story.extend(self._create_appendix(report_data))
            
            # Build PDF
            doc.build(story)
            
            # Get PDF bytes
            pdf_bytes = self.buffer.getvalue()
            self.buffer.seek(0)
            
            logger.info(f"📄 PDF report generated: {len(pdf_bytes)} bytes")
            return pdf_bytes
            
        except Exception as e:
            logger.error(f"PDF generation failed: {str(e)}")
            raise
    
    def _create_title_page(self, report_data: ReportData) -> List:
        """Create title page elements"""
        
        elements = []
        
        # Title
        elements.append(Paragraph(self.config.title, self.styles['CustomTitle']))
        elements.append(Spacer(1, 12))
        
        # Subtitle
        if self.config.subtitle:
            elements.append(Paragraph(self.config.subtitle, self.styles['CustomSubtitle']))
            elements.append(Spacer(1, 20))
        
        # Website info
        website_info = [
            f"<b>Website:</b> {report_data.website_info.get('url', 'N/A')}",
            f"<b>Unternehmen:</b> {report_data.user_info.get('company_name', 'N/A')}",
            f"<b>Berichtsdatum:</b> {report_data.generated_at.strftime('%d.%m.%Y %H:%M')}",
            f"<b>Scan-ID:</b> {report_data.scan_results.get('scan_id', 'N/A')}"
        ]
        
        for info in website_info:
            elements.append(Paragraph(info, self.styles['Normal']))
            elements.append(Spacer(1, 6))
        
        elements.append(Spacer(1, 40))
        
        # Compliance score box
        score = report_data.scan_results.get('overall_score', 0)
        score_color = self._get_score_color(score)
        
        score_text = f"""
        <para align="center">
            <b><font size="20" color="{score_color}">Compliance-Score: {score}%</font></b><br/>
            <font size="12">Gesamtbewertung der Website-Compliance</font>
        </para>
        """
        
        elements.append(Paragraph(score_text, self.styles['Normal']))
        elements.append(Spacer(1, 30))
        
        # Risk assessment
        risk = report_data.scan_results.get('total_risk_euro', 0)
        risk_text = f"""
        <para align="center">
            <b><font size="14" color="red">Geschätztes Risiko: {risk:,.0f} €</font></b><br/>
            <font size="10">Potentielle Bußgelder bei Abmahnungen</font>
        </para>
        """
        
        elements.append(Paragraph(risk_text, self.styles['Normal']))
        
        return elements
    
    def _create_executive_summary(self, report_data: ReportData) -> List:
        """Create executive summary section"""
        
        elements = []
        
        elements.append(Paragraph("Executive Summary", self.styles['SectionHeader']))
        elements.append(Spacer(1, 12))
        
        # Summary stats
        stats = report_data.compliance_metrics
        
        summary_text = f"""
        Diese Compliance-Analyse für <b>{report_data.website_info.get('url', 'N/A')}</b> 
        zeigt einen Gesamtscore von <b>{stats.get('overall_score', 0):.1f}%</b>.
        
        <br/><br/>
        
        <b>Wichtigste Erkenntnisse:</b><br/>
        • {stats.get('critical_issues', 0)} kritische Compliance-Probleme identifiziert<br/>
        • {stats.get('warning_issues', 0)} Verbesserungsmöglichkeiten gefunden<br/>
        • Geschätztes Bußgeldrisiko: {stats.get('total_risk_euro', 0):,.0f} €<br/>
        • {len([issue for issue in report_data.issues if issue.get('is_auto_fixable')])} Probleme automatisch behebbar
        
        <br/><br/>
        
        <b>Empfohlene Maßnahmen:</b><br/>
        • Sofortige Behebung aller kritischen Compliance-Probleme<br/>
        • Implementierung DSGVO-konformer Datenschutzerklärung<br/>
        • Einführung TTDSG-konformen Cookie-Banners<br/>
        • Verbesserung der Barrierefreiheit nach WCAG 2.1 AA
        """
        
        elements.append(Paragraph(summary_text, self.styles['Normal']))
        
        return elements
    
    def _create_compliance_overview(self, report_data: ReportData) -> List:
        """Create compliance overview with charts"""
        
        elements = []
        
        elements.append(Paragraph("Compliance-Übersicht", self.styles['SectionHeader']))
        elements.append(Spacer(1, 12))
        
        # Create compliance table
        results = report_data.scan_results.get('results', [])
        
        table_data = [['Kategorie', 'Score', 'Status', 'Risiko (€)', 'Empfehlung']]
        
        for result in results:
            status_icon = "✅" if result.get('status') == 'pass' else "⚠️" if result.get('status') == 'warning' else "❌"
            table_data.append([
                result.get('category', 'N/A'),
                f"{result.get('score', 0)}%",
                status_icon,
                f"{result.get('risk_euro', 0):,.0f}",
                "Sofort beheben" if result.get('status') == 'fail' else "Überwachen"
            ])
        
        table = Table(table_data, colWidths=[2.5*inch, 0.8*inch, 0.6*inch, 1*inch, 1.5*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498DB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 20))
        
        return elements
    
    def _create_issues_section(self, report_data: ReportData) -> List:
        """Create detailed issues section"""
        
        elements = []
        
        elements.append(Paragraph("Identifizierte Probleme", self.styles['SectionHeader']))
        elements.append(Spacer(1, 12))
        
        # Group issues by severity
        critical_issues = [issue for issue in report_data.issues if issue.get('severity') == 'critical']
        warning_issues = [issue for issue in report_data.issues if issue.get('severity') == 'warning']
        
        # Critical issues
        if critical_issues:
            elements.append(Paragraph("🚨 Kritische Probleme", self.styles['Heading3']))
            elements.append(Spacer(1, 6))
            
            for issue in critical_issues:
                issue_text = f"""
                <b>{issue.get('title', 'Unbekanntes Problem')}</b><br/>
                {issue.get('description', 'Keine Beschreibung verfügbar')}<br/>
                <i>Rechtliche Grundlage:</i> {issue.get('legal_basis', 'N/A')}<br/>
                <i>Risiko:</i> {issue.get('risk_euro', 0):,.0f} €
                """
                
                elements.append(Paragraph(issue_text, self.styles['CriticalIssue']))
                elements.append(Spacer(1, 8))
        
        # Warning issues  
        if warning_issues:
            elements.append(Paragraph("⚠️ Verbesserungsmöglichkeiten", self.styles['Heading3']))
            elements.append(Spacer(1, 6))
            
            for issue in warning_issues:
                issue_text = f"""
                <b>{issue.get('title', 'Unbekanntes Problem')}</b><br/>
                {issue.get('description', 'Keine Beschreibung verfügbar')}<br/>
                <i>Empfehlung:</i> {issue.get('recommendation', 'Keine Empfehlung verfügbar')}
                """
                
                elements.append(Paragraph(issue_text, self.styles['Normal']))
                elements.append(Spacer(1, 8))
        
        return elements
    
    def _create_technical_section(self, report_data: ReportData) -> List:
        """Create technical analysis section"""
        
        elements = []
        
        elements.append(Paragraph("Technische Analyse", self.styles['SectionHeader']))
        elements.append(Spacer(1, 12))
        
        tech_data = report_data.technical_analysis or {}
        
        tech_text = f"""
        <b>Website-Performance:</b><br/>
        • Ladezeit: {tech_data.get('page_load_time', 'N/A')} Sekunden<br/>
        • Seitengröße: {tech_data.get('page_size', 'N/A')} MB<br/>
        • HTTPS aktiviert: {'✅ Ja' if tech_data.get('https_enabled') else '❌ Nein'}<br/>
        
        <br/>
        
        <b>Cookies & Tracking:</b><br/>
        • Erkannte Cookies: {tech_data.get('cookies_detected', 0)}<br/>
        • Tracking-Domains: {len(tech_data.get('tracking_domains', []))}<br/>
        • Third-Party Services: {len(tech_data.get('third_party_integrations', []))}<br/>
        
        <br/>
        
        <b>Barrierefreiheit:</b><br/>
        • Bilder ohne Alt-Text: {tech_data.get('images_without_alt', 0)}<br/>
        • Formulare gefunden: {tech_data.get('forms_found', 0)}<br/>
        • Überschriftenstruktur: {'✅ Korrekt' if tech_data.get('heading_structure_valid') else '⚠️ Verbesserungsbedarf'}
        """
        
        elements.append(Paragraph(tech_text, self.styles['Normal']))
        
        return elements
    
    def _create_recommendations_section(self, report_data: ReportData) -> List:
        """Create recommendations section"""
        
        elements = []
        
        elements.append(Paragraph("Handlungsempfehlungen", self.styles['SectionHeader']))
        elements.append(Spacer(1, 12))
        
        for i, recommendation in enumerate(report_data.recommendations, 1):
            rec_text = f"""
            <b>{i}. {recommendation.get('title', 'Empfehlung')}</b><br/>
            {recommendation.get('description', 'Keine Beschreibung verfügbar')}<br/>
            <i>Priorität:</i> {recommendation.get('priority', 'Medium')}<br/>
            <i>Geschätzter Aufwand:</i> {recommendation.get('effort', 'Unbekannt')}
            """
            
            elements.append(Paragraph(rec_text, self.styles['Recommendation']))
            elements.append(Spacer(1, 10))
        
        return elements
    
    def _create_appendix(self, report_data: ReportData) -> List:
        """Create appendix section"""
        
        elements = []
        
        elements.append(PageBreak())
        elements.append(Paragraph("Anhang", self.styles['SectionHeader']))
        elements.append(Spacer(1, 12))
        
        # Legal information
        legal_text = """
        <b>Rechtliche Hinweise:</b><br/>
        
        Diese Compliance-Analyse wurde automatisiert durch die Complyo-Plattform erstellt 
        und stellt eine Momentaufnahme zum angegebenen Zeitpunkt dar. Die Ergebnisse ersetzen 
        keine individuelle Rechtsberatung.<br/><br/>
        
        <b>Relevante Gesetze und Verordnungen:</b><br/>
        • DSGVO (Datenschutz-Grundverordnung)<br/>
        • TTDSG (Telekommunikation-Telemedien-Datenschutz-Gesetz)<br/>  
        • TMG (Telemediengesetz)<br/>
        • WCAG 2.1 AA (Web Content Accessibility Guidelines)<br/>
        • BITV 2.0 (Barrierefreie-Informationstechnik-Verordnung)<br/><br/>
        
        <b>Über Complyo:</b><br/>
        Complyo ist eine professionelle Plattform für Website-Compliance-Management. 
        Wir helfen Unternehmen dabei, ihre Websites rechtssicher zu gestalten und 
        Compliance-Risiken zu minimieren.<br/><br/>
        
        Für weitere Informationen besuchen Sie uns unter: https://complyo.tech
        """
        
        elements.append(Paragraph(legal_text, self.styles['Normal']))
        
        return elements
    
    def _get_score_color(self, score: float) -> str:
        """Get color based on compliance score"""
        
        if score >= 80:
            return "#27AE60"  # Green
        elif score >= 60:
            return "#F39C12"  # Orange
        else:
            return "#E74C3C"  # Red

class ExcelReportGenerator:
    """Professional Excel report generator"""
    
    def __init__(self, config: ReportConfig):
        """Initialize Excel generator with configuration"""
        
        self.config = config
        self.workbook = openpyxl.Workbook()
        
        # Remove default sheet
        default_sheet = self.workbook.active
        self.workbook.remove(default_sheet)
        
        # Define styles
        self._setup_styles()
    
    def _setup_styles(self):
        """Setup Excel styles"""
        
        # Header style
        self.header_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        
        # Subheader style
        self.subheader_font = Font(name='Arial', size=12, bold=True, color='2C3E50')
        self.subheader_fill = PatternFill(start_color='ECF0F1', end_color='ECF0F1', fill_type='solid')
        
        # Critical issue style
        self.critical_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        self.critical_fill = PatternFill(start_color='E74C3C', end_color='E74C3C', fill_type='solid')
        
        # Warning style
        self.warning_font = Font(name='Arial', size=10, color='D68910')
        self.warning_fill = PatternFill(start_color='FEF9E7', end_color='FEF9E7', fill_type='solid')
        
        # Success style
        self.success_font = Font(name='Arial', size=10, color='27AE60')
        self.success_fill = PatternFill(start_color='D5F4E6', end_color='D5F4E6', fill_type='solid')
        
        # Border
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    async def generate_compliance_report(self, report_data: ReportData) -> bytes:
        """Generate comprehensive compliance report Excel"""
        
        try:
            # Create sheets
            self._create_overview_sheet(report_data)
            self._create_issues_sheet(report_data)
            self._create_technical_sheet(report_data)
            self._create_recommendations_sheet(report_data)
            
            if report_data.monitoring_data:
                self._create_monitoring_sheet(report_data)
            
            # Save to buffer
            buffer = io.BytesIO()
            self.workbook.save(buffer)
            excel_bytes = buffer.getvalue()
            buffer.close()
            
            logger.info(f"📊 Excel report generated: {len(excel_bytes)} bytes")
            return excel_bytes
            
        except Exception as e:
            logger.error(f"Excel generation failed: {str(e)}")
            raise
    
    def _create_overview_sheet(self, report_data: ReportData):
        """Create overview sheet"""
        
        ws = self.workbook.create_sheet("Compliance Übersicht")
        
        # Title
        ws['A1'] = self.config.title
        ws['A1'].font = Font(name='Arial', size=16, bold=True, color='2C3E50')
        ws.merge_cells('A1:E1')
        
        # Basic info
        row = 3
        info_data = [
            ['Website', report_data.website_info.get('url', 'N/A')],
            ['Unternehmen', report_data.user_info.get('company_name', 'N/A')],
            ['Berichtsdatum', report_data.generated_at.strftime('%d.%m.%Y %H:%M')],
            ['Scan-ID', report_data.scan_results.get('scan_id', 'N/A')]
        ]
        
        for label, value in info_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = self.subheader_font
            ws[f'B{row}'] = value
            row += 1
        
        # Compliance metrics
        row += 2
        ws[f'A{row}'] = 'Compliance-Metriken'
        ws[f'A{row}'].font = self.header_font
        ws[f'A{row}'].fill = self.header_fill
        ws.merge_cells(f'A{row}:E{row}')
        
        row += 1
        metrics_data = [
            ['Gesamtscore', f"{report_data.scan_results.get('overall_score', 0):.1f}%"],
            ['Kritische Probleme', report_data.scan_results.get('critical_issues', 0)],
            ['Warnungen', report_data.scan_results.get('warning_issues', 0)],
            ['Geschätztes Risiko', f"{report_data.scan_results.get('total_risk_euro', 0):,.0f} €"]
        ]
        
        for label, value in metrics_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
        
        # Compliance results table
        row += 2
        ws[f'A{row}'] = 'Kategorie'
        ws[f'B{row}'] = 'Score'
        ws[f'C{row}'] = 'Status'
        ws[f'D{row}'] = 'Risiko (€)'
        ws[f'E{row}'] = 'Beschreibung'
        
        # Style header row
        for col in ['A', 'B', 'C', 'D', 'E']:
            cell = ws[f'{col}{row}']
            cell.font = self.header_font
            cell.fill = self.header_fill
        
        row += 1
        
        # Add results data
        results = report_data.scan_results.get('results', [])
        for result in results:
            ws[f'A{row}'] = result.get('category', 'N/A')
            ws[f'B{row}'] = f"{result.get('score', 0)}%"
            
            status = result.get('status', 'unknown')
            ws[f'C{row}'] = 'Bestanden' if status == 'pass' else 'Warnung' if status == 'warning' else 'Fehler'
            
            # Style status cell
            if status == 'pass':
                ws[f'C{row}'].font = self.success_font
                ws[f'C{row}'].fill = self.success_fill
            elif status == 'warning':
                ws[f'C{row}'].font = self.warning_font
                ws[f'C{row}'].fill = self.warning_fill
            else:
                ws[f'C{row}'].font = self.critical_font
                ws[f'C{row}'].fill = self.critical_fill
            
            ws[f'D{row}'] = result.get('risk_euro', 0)
            ws[f'E{row}'] = result.get('description', 'Keine Beschreibung')[:50] + '...'
            
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    def _create_issues_sheet(self, report_data: ReportData):
        """Create issues detail sheet"""
        
        ws = self.workbook.create_sheet("Probleme Details")
        
        # Headers
        headers = ['Kategorie', 'Schweregrad', 'Titel', 'Beschreibung', 'Rechtliche Grundlage', 'Risiko (€)', 'Auto-Fix', 'Empfehlung']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
        
        # Add issues data
        row = 2
        for issue in report_data.issues:
            ws.cell(row=row, column=1, value=issue.get('category', 'N/A'))
            
            severity = issue.get('severity', 'unknown')
            severity_cell = ws.cell(row=row, column=2, value=severity.title())
            
            # Style severity cell
            if severity == 'critical':
                severity_cell.font = self.critical_font
                severity_cell.fill = self.critical_fill
            elif severity == 'warning':
                severity_cell.font = self.warning_font
                severity_cell.fill = self.warning_fill
            
            ws.cell(row=row, column=3, value=issue.get('title', 'N/A'))
            ws.cell(row=row, column=4, value=issue.get('description', 'N/A'))
            ws.cell(row=row, column=5, value=issue.get('legal_basis', 'N/A'))
            ws.cell(row=row, column=6, value=issue.get('risk_euro', 0))
            ws.cell(row=row, column=7, value='Ja' if issue.get('is_auto_fixable') else 'Nein')
            ws.cell(row=row, column=8, value=issue.get('recommendation', 'N/A'))
            
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 80)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    def _create_technical_sheet(self, report_data: ReportData):
        """Create technical analysis sheet"""
        
        ws = self.workbook.create_sheet("Technische Analyse")
        
        if not report_data.technical_analysis:
            ws['A1'] = 'Keine technischen Daten verfügbar'
            return
        
        tech_data = report_data.technical_analysis
        
        # Performance metrics
        ws['A1'] = 'Performance Metriken'
        ws['A1'].font = self.header_font
        ws['A1'].fill = self.header_fill
        
        row = 2
        perf_data = [
            ['Ladezeit (Sekunden)', tech_data.get('page_load_time', 'N/A')],
            ['Seitengröße (MB)', tech_data.get('page_size', 'N/A')],
            ['HTTPS aktiviert', 'Ja' if tech_data.get('https_enabled') else 'Nein'],
            ['Response Code', tech_data.get('status_code', 'N/A')]
        ]
        
        for label, value in perf_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
        
        # Cookies & Tracking
        row += 2
        ws[f'A{row}'] = 'Cookies & Tracking'
        ws[f'A{row}'].font = self.header_font
        ws[f'A{row}'].fill = self.header_fill
        
        row += 1
        cookie_data = [
            ['Erkannte Cookies', tech_data.get('cookies_detected', 0)],
            ['Tracking Domains', len(tech_data.get('tracking_domains', []))],
            ['Third-Party Services', len(tech_data.get('third_party_integrations', []))]
        ]
        
        for label, value in cookie_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
        
        # List tracking domains
        if tech_data.get('tracking_domains'):
            row += 2
            ws[f'A{row}'] = 'Tracking Domains:'
            ws[f'A{row}'].font = self.subheader_font
            row += 1
            
            for domain in tech_data['tracking_domains']:
                ws[f'A{row}'] = domain
                row += 1
    
    def _create_recommendations_sheet(self, report_data: ReportData):
        """Create recommendations sheet"""
        
        ws = self.workbook.create_sheet("Handlungsempfehlungen")
        
        # Headers
        headers = ['Priorität', 'Titel', 'Beschreibung', 'Kategorie', 'Aufwand', 'Auto-Fix möglich']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
        
        # Add recommendations
        row = 2
        for recommendation in report_data.recommendations:
            priority = recommendation.get('priority', 'medium')
            priority_cell = ws.cell(row=row, column=1, value=priority.title())
            
            # Style priority cell
            if priority == 'high':
                priority_cell.font = self.critical_font
                priority_cell.fill = self.critical_fill
            elif priority == 'medium':
                priority_cell.font = self.warning_font
                priority_cell.fill = self.warning_fill
            
            ws.cell(row=row, column=2, value=recommendation.get('title', 'N/A'))
            ws.cell(row=row, column=3, value=recommendation.get('description', 'N/A'))
            ws.cell(row=row, column=4, value=recommendation.get('category', 'N/A'))
            ws.cell(row=row, column=5, value=recommendation.get('effort', 'N/A'))
            ws.cell(row=row, column=6, value='Ja' if recommendation.get('auto_fixable') else 'Nein')
            
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 60)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    def _create_monitoring_sheet(self, report_data: ReportData):
        """Create monitoring data sheet"""
        
        if not report_data.monitoring_data:
            return
        
        ws = self.workbook.create_sheet("Monitoring Verlauf")
        
        # Headers for monitoring history
        headers = ['Datum', 'Score', 'Kritische Probleme', 'Warnungen', 'Risiko (€)', 'Status']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
        
        # Add monitoring data (if available)
        monitoring_history = report_data.monitoring_data.get('scan_history', [])
        
        row = 2
        for scan in monitoring_history:
            ws.cell(row=row, column=1, value=scan.get('date', 'N/A'))
            ws.cell(row=row, column=2, value=f"{scan.get('score', 0):.1f}%")
            ws.cell(row=row, column=3, value=scan.get('critical_issues', 0))
            ws.cell(row=row, column=4, value=scan.get('warnings', 0))
            ws.cell(row=row, column=5, value=scan.get('risk_euro', 0))
            ws.cell(row=row, column=6, value=scan.get('status', 'N/A'))
            row += 1

class ReportService:
    """Main report generation service"""
    
    def __init__(self):
        """Initialize report service"""
        
        self.temp_dir = tempfile.gettempdir()
        logger.info("📄 Report service initialized")
    
    async def generate_pdf_report(self, report_data: ReportData, config: ReportConfig = None) -> bytes:
        """Generate PDF compliance report"""
        
        if config is None:
            config = ReportConfig(
                title="Website Compliance Report",
                subtitle="Automatisierte DSGVO & Rechts-Compliance Analyse"
            )
        
        generator = PDFReportGenerator(config)
        return await generator.generate_compliance_report(report_data)
    
    async def generate_excel_report(self, report_data: ReportData, config: ReportConfig = None) -> bytes:
        """Generate Excel compliance report"""
        
        if config is None:
            config = ReportConfig(
                title="Website Compliance Report",
                subtitle="Detaillierte Compliance-Analyse"
            )
        
        generator = ExcelReportGenerator(config)
        return await generator.generate_compliance_report(report_data)
    
    async def generate_both_reports(self, report_data: ReportData, config: ReportConfig = None) -> Dict[str, bytes]:
        """Generate both PDF and Excel reports"""
        
        pdf_bytes, excel_bytes = await asyncio.gather(
            self.generate_pdf_report(report_data, config),
            self.generate_excel_report(report_data, config)
        )
        
        return {
            "pdf": pdf_bytes,
            "excel": excel_bytes
        }
    
    def create_report_data_from_scan(self, scan_result: Dict[str, Any], user_info: Dict[str, Any]) -> ReportData:
        """Create ReportData from scan result"""
        
        # Extract website info
        website_info = {
            "url": scan_result.get("url", ""),
            "name": scan_result.get("url", "").replace("https://", "").replace("http://", "")
        }
        
        # Extract compliance metrics
        compliance_metrics = {
            "overall_score": scan_result.get("overall_score", 0),
            "total_issues": scan_result.get("total_issues", 0),
            "critical_issues": scan_result.get("critical_issues", 0),
            "warning_issues": scan_result.get("warning_issues", 0),
            "total_risk_euro": scan_result.get("total_risk_euro", 0)
        }
        
        # Convert results to issues format
        issues = []
        for result in scan_result.get("results", []):
            issues.append({
                "category": result.get("category", ""),
                "title": result.get("message", ""),
                "description": result.get("description", ""),
                "severity": "critical" if result.get("status") == "fail" else "warning",
                "legal_basis": result.get("legal_basis", ""),
                "risk_euro": result.get("risk_euro", 0),
                "is_auto_fixable": result.get("auto_fixable", False),
                "recommendation": result.get("recommendation", "")
            })
        
        # Create recommendations from next steps
        recommendations = []
        for step in scan_result.get("next_steps", []):
            recommendations.append({
                "title": step.get("title", ""),
                "description": step.get("description", ""),
                "category": step.get("action", ""),
                "priority": "high" if step.get("action") == "expert_service" else "medium",
                "effort": "Gering" if step.get("action") == "ai_fix" else "Hoch",
                "auto_fixable": step.get("action") == "ai_fix"
            })
        
        return ReportData(
            scan_results=scan_result,
            user_info=user_info,
            website_info=website_info,
            compliance_metrics=compliance_metrics,
            issues=issues,
            recommendations=recommendations,
            technical_analysis=scan_result.get("technical_analysis")
        )

# Global report service instance
report_service = ReportService()